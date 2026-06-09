"""Excel export — period reports and Purchase Invoices Database."""

from __future__ import annotations

import io
from datetime import date, datetime
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from core.debug_logger import debug_trace, get_logger
from schemas.invoice import InvoiceResponse
from schemas.report import PeriodReportResponse

logger = get_logger(__name__)

_COLOR_NAVY_HEADER = "0D123F"
_COLOR_HEADER_TEXT = "FFFFFF"
_COLOR_ROW_ALT = "F0F2F8"
_COLOR_ROW_EVEN = "FFFFFF"
_COLOR_TEXT = "1A1F4D"
_COLOR_BORDER = "C8CDDF"

_FMT_DATE = "DD/MM/YYYY"
_FMT_AMOUNT = "#,##0.00"

_FONT_HEADER = Font(name="Calibri", bold=True, color=_COLOR_HEADER_TEXT, size=11)
_FONT_BODY = Font(name="Calibri", color=_COLOR_TEXT, size=11)

_FILL_HEADER = PatternFill("solid", fgColor=_COLOR_NAVY_HEADER)
_FILL_ROW_ALT = PatternFill("solid", fgColor=_COLOR_ROW_ALT)
_FILL_ROW_EVEN = PatternFill("solid", fgColor=_COLOR_ROW_EVEN)

_ALIGN_HEADER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_DATE = Alignment(horizontal="center", vertical="center")
_ALIGN_AMOUNT = Alignment(horizontal="right", vertical="center")
_ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
_ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

_PURCHASE_INVOICE_HEADERS: tuple[str, ...] = (
    "Invoice Date",
    "Name of Company",
    "Adress of Company",
    "Invoice Number",
    "Amount",
    "Debt",
    "Account Details",
    "Internal Note/Description",
    "Client / Employee Related",
    "paid at (Date)",
    "paid by",
    "Fixed/Not fixed",
    "Category",
)

_PURCHASE_INVOICE_COL_WIDTHS: tuple[float, ...] = (
    13.0,
    32.0,
    34.0,
    22.0,
    14.0,
    12.0,
    38.0,
    42.0,
    24.0,
    14.0,
    18.0,
    16.0,
    22.0,
)

# 1-based column indices for body formatting
_PI_COL_DATES = frozenset({1, 10})
_PI_COL_AMOUNTS = frozenset({5, 6})
_PI_COL_INVOICE_NUMBER = 4
_PI_COL_TEXT_LEFT = frozenset({2, 3, 7, 8, 9, 11, 12, 13})

_THIN_BORDER = Border(
    left=Side(style="thin", color=_COLOR_BORDER),
    right=Side(style="thin", color=_COLOR_BORDER),
    top=Side(style="thin", color=_COLOR_BORDER),
    bottom=Side(style="thin", color=_COLOR_BORDER),
)

_COL_DATES = frozenset({2, 3})
_COL_AMOUNTS = frozenset({2, 3})


def _apply_cell_format(cell, *, is_header: bool, row_fill: PatternFill) -> None:
    cell.border = _THIN_BORDER
    cell.fill = _FILL_HEADER if is_header else row_fill
    cell.font = _FONT_HEADER if is_header else _FONT_BODY
    cell.alignment = _ALIGN_HEADER if is_header else _ALIGN_CENTER


def _excel_safe(value: object | None) -> object | None:
    if not isinstance(value, str) or not value:
        return value
    if value[0] in ("=", "+", "-", "@", "\t", "\r"):
        return "'" + value
    return value


def _decimal_to_float(value: Decimal | float | int | None) -> float | None:
    if value is None:
        return None
    return float(value)


def _invoice_to_export_row(inv: InvoiceResponse) -> list[object | None]:
    invoice_date = inv.invoice_date
    if isinstance(invoice_date, date) and not isinstance(invoice_date, datetime):
        invoice_date = datetime(
            invoice_date.year, invoice_date.month, invoice_date.day
        )
    paid_at = inv.paid_at_date
    if isinstance(paid_at, date) and not isinstance(paid_at, datetime):
        paid_at = datetime(paid_at.year, paid_at.month, paid_at.day)
    return [
        invoice_date,
        _excel_safe(inv.name_of_company),
        _excel_safe(inv.address_of_company),
        _excel_safe(
            str(inv.invoice_number) if inv.invoice_number is not None else None
        ),
        _decimal_to_float(inv.amount),
        _decimal_to_float(inv.debt),
        _excel_safe(inv.account_details),
        _excel_safe(inv.internal_note_description),
        _excel_safe(inv.client_employee_related),
        paid_at,
        _excel_safe(inv.paid_by),
        _excel_safe(inv.fixed_status),
        _excel_safe(inv.category),
    ]


def _apply_purchase_invoice_body_cell(cell, *, col_idx: int, row_fill: PatternFill) -> None:
    cell.border = _THIN_BORDER
    cell.fill = row_fill
    cell.font = _FONT_BODY
    if col_idx in _PI_COL_DATES and isinstance(cell.value, datetime):
        cell.number_format = _FMT_DATE
        cell.alignment = _ALIGN_DATE
    elif col_idx in _PI_COL_AMOUNTS and isinstance(cell.value, (int, float)):
        cell.number_format = _FMT_AMOUNT
        cell.alignment = _ALIGN_AMOUNT
    elif col_idx == _PI_COL_INVOICE_NUMBER:
        cell.number_format = "@"
        cell.alignment = _ALIGN_CENTER
    elif col_idx in _PI_COL_TEXT_LEFT:
        cell.alignment = _ALIGN_LEFT
    else:
        cell.alignment = _ALIGN_CENTER


class ExcelService:
    @debug_trace
    def write_purchase_invoices_workbook(
        self, invoices: list[InvoiceResponse]
    ) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Purchase Invoices"
        ws.append(list(_PURCHASE_INVOICE_HEADERS))

        for inv in invoices:
            ws.append(_invoice_to_export_row(inv))

        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False

        for col_idx, width in enumerate(_PURCHASE_INVOICE_COL_WIDTHS, start=1):
            letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[letter].width = width

        for row_idx in range(1, ws.max_row + 1):
            is_header = row_idx == 1
            row_fill = (
                _FILL_ROW_EVEN if is_header or row_idx % 2 == 0 else _FILL_ROW_ALT
            )
            for col_idx in range(1, len(_PURCHASE_INVOICE_HEADERS) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = _THIN_BORDER
                cell.fill = _FILL_HEADER if is_header else row_fill
                cell.font = _FONT_HEADER if is_header else _FONT_BODY
                if is_header:
                    cell.alignment = _ALIGN_HEADER
                else:
                    _apply_purchase_invoice_body_cell(
                        cell, col_idx=col_idx, row_fill=row_fill
                    )

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    @debug_trace
    def write_period_report_workbook(self, report: PeriodReportResponse) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"

        summary_rows = [
            ("Period", report.period_label),
            ("Start date", report.start_date),
            ("End date", report.end_date),
            ("Total invoices", report.total_invoices),
            ("Total amount", float(report.total_amount)),
            ("Paid invoices", report.paid_invoices),
            ("Unpaid invoices", report.unpaid_invoices),
            ("Total paid amount", float(report.total_paid_amount)),
            ("Matched invoices", report.matched_invoices),
            ("Unmatched invoices", report.unmatched_invoices),
            ("Needs review (invoices)", report.needs_review),
            ("Bank transactions", report.bank_transactions),
            ("Bank matched", report.bank_matched),
            ("Bank needs review", report.bank_needs_review),
        ]

        ws.append(["Metric", "Value"])
        for label, value in summary_rows:
            ws.append([label, value])

        cat_ws = wb.create_sheet("By category")
        cat_ws.append(["Category", "Count", "Total amount"])
        for row in report.by_category:
            cat_ws.append([row.category, row.count, float(row.total_amount)])

        for sheet in (ws, cat_ws):
            sheet.sheet_view.showGridLines = False
            sheet.column_dimensions["A"].width = 28
            sheet.column_dimensions["B"].width = 18
            if sheet.title == "By category":
                sheet.column_dimensions["C"].width = 16
            for row_idx in range(1, sheet.max_row + 1):
                is_header = row_idx == 1
                row_fill = (
                    _FILL_ROW_EVEN if is_header or row_idx % 2 == 0 else _FILL_ROW_ALT
                )
                for col_idx in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    _apply_cell_format(
                        cell,
                        is_header=is_header,
                        row_fill=row_fill,
                    )
                    if not is_header and col_idx in _COL_DATES and isinstance(
                        cell.value, date
                    ):
                        cell.number_format = _FMT_DATE
                        cell.alignment = _ALIGN_DATE
                    if not is_header and col_idx in _COL_AMOUNTS and isinstance(
                        cell.value, (int, float)
                    ):
                        cell.number_format = _FMT_AMOUNT
                        cell.alignment = _ALIGN_AMOUNT

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
