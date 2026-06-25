"""Purchase Invoices Database Excel export layout."""

from datetime import date, datetime, timezone
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook

from schemas.invoice import InvoiceResponse
from services.excel_service import (
    ExcelService,
    _PURCHASE_INVOICE_HEADERS,
    _sort_purchase_invoices_for_export,
)

EXPECTED_HEADERS = (
    "Invoice Date",
    "Name of Company",
    "Adress of Company",
    "Invoice Number",
    "Amount",
    "Debt",
    "Account Details",
    "Internal Note/Description",
    "Client / Employee Related",
    "paid at (Date)",
    "paid by",
    "Fixed/Not fixed",
    "Category",
)


def _sample_invoice() -> InvoiceResponse:
    return InvoiceResponse(
        id=1,
        invoice_date=date(2026, 2, 23),
        name_of_company="SCHMIEDER it-solutions GmbH",
        address_of_company="Carl-Zeiss-Straße 5",
        invoice_number="613260192",
        amount=Decimal("2649.24"),
        debt=None,
        currency="EUR",
        original_amount=Decimal("2649.24"),
        original_currency="EUR",
        exchange_rate=Decimal("1"),
        exchange_rate_date=date(2026, 2, 23),
        account_details="IBAN DE29 6035 0130 0001 5990 60",
        internal_note_description="Microsoft 365 services",
        client_employee_related="Muavi Rexhepi",
        paid_at_date=None,
        paid_by=None,
        fixed_status=None,
        category="Software",
        extraction_confidence=Decimal("0.95"),
        field_confidences=None,
        review_reasons=None,
        review_status="approved",
        match_status="unmatched",
        source_file_id=1,
        uploaded_by=1,
        created_at=datetime(2026, 2, 23, tzinfo=timezone.utc),
        updated_at=datetime(2026, 2, 23, tzinfo=timezone.utc),
    )


def _sample_invoice_with(
    *,
    invoice_id: int,
    invoice_date: date | None,
) -> InvoiceResponse:
    base = _sample_invoice()
    return base.model_copy(update={"id": invoice_id, "invoice_date": invoice_date})


def test_purchase_invoices_export_sorts_by_paid_date_not_upload_id():
    paid_dates = {
        40: date(2026, 6, 3),
        39: date(2026, 6, 10),
        1: date(2026, 6, 25),
        16: date(2026, 6, 15),
    }
    invoices = [
        _sample_invoice_with(invoice_id=40, invoice_date=date(2026, 2, 1)).model_copy(
            update={"paid_at_date": paid_dates[40]}
        ),
        _sample_invoice_with(invoice_id=39, invoice_date=date(2026, 2, 4)).model_copy(
            update={"paid_at_date": paid_dates[39]}
        ),
        _sample_invoice_with(invoice_id=1, invoice_date=date(2026, 2, 20)).model_copy(
            update={"paid_at_date": paid_dates[1]}
        ),
        _sample_invoice_with(invoice_id=16, invoice_date=date(2026, 2, 9)).model_copy(
            update={"paid_at_date": paid_dates[16]}
        ),
    ]
    ordered = _sort_purchase_invoices_for_export(invoices)
    assert [inv.id for inv in ordered] == [1, 16, 39, 40]


def test_purchase_invoices_export_sorts_by_invoice_date_not_upload_id():
    invoices = [
        _sample_invoice_with(invoice_id=40, invoice_date=date(2026, 2, 1)),
        _sample_invoice_with(invoice_id=39, invoice_date=date(2026, 2, 4)),
        _sample_invoice_with(invoice_id=1, invoice_date=date(2026, 2, 20)),
        _sample_invoice_with(invoice_id=16, invoice_date=date(2026, 2, 9)),
    ]
    ordered = _sort_purchase_invoices_for_export(
        invoices, sort="invoice_date_desc"
    )
    assert [inv.id for inv in ordered] == [1, 16, 39, 40]
    assert [inv.invoice_date for inv in ordered] == [
        date(2026, 2, 20),
        date(2026, 2, 9),
        date(2026, 2, 4),
        date(2026, 2, 1),
    ]


def test_purchase_invoices_workbook_matches_official_columns():
    data = ExcelService().write_purchase_invoices_workbook([_sample_invoice()])
    wb = load_workbook(BytesIO(data), read_only=True)
    ws = wb.active

    assert ws.title == "Purchase Invoices"
    assert tuple(_PURCHASE_INVOICE_HEADERS) == EXPECTED_HEADERS
    assert [c.value for c in ws[1]] == list(EXPECTED_HEADERS)

    row = [c.value for c in ws[2]]
    assert row[1] == "SCHMIEDER it-solutions GmbH"
    assert row[3] == "613260192"
    assert row[4] == 2649.24
    assert row[12] == "Software"

    wb.close()
