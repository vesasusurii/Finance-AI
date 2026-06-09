from datetime import date, datetime, timezone
from decimal import Decimal

from openpyxl import load_workbook

from schemas.invoice import InvoiceResponse
from services.excel_service import ExcelService


def _sample_invoice(**overrides) -> InvoiceResponse:
    data = {
        "id": 1,
        "invoice_number": "=CMD('calc')",
        "name_of_company": "+evil",
        "address_of_company": "Addr",
        "amount": Decimal("10.00"),
        "debt": Decimal("0"),
        "currency": "EUR",
        "original_amount": None,
        "original_currency": None,
        "exchange_rate": None,
        "exchange_rate_date": None,
        "invoice_date": date(2026, 1, 15),
        "account_details": None,
        "internal_note_description": None,
        "client_employee_related": None,
        "paid_at_date": None,
        "paid_by": None,
        "fixed_status": None,
        "category": "Services",
        "review_status": "approved",
        "match_status": "unmatched",
        "uploaded_by": 1,
        "extraction_confidence": 0.9,
        "field_confidences": None,
        "review_reasons": None,
        "source_file_id": 1,
        "created_at": datetime(2026, 1, 15, tzinfo=timezone.utc),
        "updated_at": datetime(2026, 1, 15, tzinfo=timezone.utc),
    }
    data.update(overrides)
    return InvoiceResponse.model_validate(data)


def test_export_prefixes_formula_like_strings():
    svc = ExcelService()
    raw = svc.write_purchase_invoices_workbook([_sample_invoice()])
    wb = load_workbook(filename=__import__("io").BytesIO(raw))
    ws = wb.active
    assert ws.cell(row=2, column=2).value == "'+evil"
    assert ws.cell(row=2, column=4).value == "'=CMD('calc')"
