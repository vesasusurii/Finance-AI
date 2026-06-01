"""Parse ProCredit-style bank statement Excel files (doc 10).

Handles Albanian (Data / Komenti / Tipi) and German (Datum / Buchungsdatum /
Valuta / Verwendungszweck) bank exports. Date cells may arrive as Python
datetime, ISO/EU strings, or raw Excel serial numbers — all are accepted.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from typing import Any

from core.debug_logger import debug_trace, get_logger
from core.exceptions import ExcelParseError
from utils.invoice_number_parser import extract_invoice_numbers

logger = get_logger(__name__)

STOP_MARKERS = ("përmbledhje", "summary", "fsdk", "deposit insurance")
HEADER_SCAN_ROWS = 15

# Substring tokens accepted in header cells (case-insensitive). Add new bank
# exports here rather than per-cell matching — one extra token covers every
# row of the file.
REQUIRED_HEADERS = {
    "date": (
        "data",        # Albanian / Italian / Spanish
        "date",        # English
        "datum",       # German / Dutch — also matches "buchungsdatum" / "valutadatum"
        "valuta",      # German "Valuta(datum)"
    ),
    "comment": (
        "komenti",            # Albanian
        "comment",            # English
        "verwendungszweck",   # German "purpose of transfer"
        "beschreibung",       # German "description"
        "betreff",            # German "subject"
        "purpose",            # English alt
    ),
}

# Excel stores dates as days since 1899-12-30 (the "-30" accounts for Lotus's
# 1900 leap-year bug that Excel preserved for compatibility).
_EXCEL_EPOCH = datetime(1899, 12, 30)

_DATE_STRING_FORMATS: tuple[str, ...] = (
    "%d.%m.%Y",
    "%d/%m/%Y",
    "%Y-%m-%d",
    "%d-%m-%Y",
    "%d.%m.%y",
    "%d/%m/%y",
    "%Y/%m/%d",
    "%d.%m.%Y %H:%M",
    "%d.%m.%Y %H:%M:%S",
    "%d/%m/%Y %H:%M",
    "%d/%m/%Y %H:%M:%S",
    "%Y-%m-%d %H:%M",
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%dT%H:%M:%S",
)

# Excel "text" dates: leading apostrophe, NBSP, zero-width marks, comma separators.
_ZERO_WIDTH_RE = re.compile(r"[\u200e\u200f\u202a-\u202e\ufeff]")
_DATE_DMY_RE = re.compile(
    r"^(?P<d>\d{1,2})[./,\-](?P<m>\d{1,2})[./,\-](?P<y>\d{2,4})"
)


@dataclass
class ParsedBankRow:
    transaction_date: date | None
    debited_amount: Decimal | None
    credited_amount: Decimal | None
    transaction_type: str | None
    comment: str | None
    detected_invoice_numbers: list[str]


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\n", " ").replace("\r", " ")
    return re.sub(r"\s+", " ", text).strip().lower()


def _cell_str(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def _parse_amount(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return Decimal(str(value)).quantize(Decimal("0.01"))
    text = str(value).strip().replace(" ", "")
    if not text:
        return None
    text = text.replace(",", ".")
    try:
        return Decimal(text).quantize(Decimal("0.01"))
    except InvalidOperation:
        return None


def _sanitize_date_text(raw: str) -> str:
    """Normalize plain-text Excel date cells before strptime / regex parsing."""
    text = _ZERO_WIDTH_RE.sub("", str(raw).strip())
    text = text.replace("\xa0", " ").strip()
    # Excel forces text dates with a leading apostrophe in the stored value.
    text = text.lstrip("'\"").rstrip("'\"")
    return text.strip()


def _expand_two_digit_year(year: int) -> int:
    if year >= 100:
        return year
    return year + 2000 if year < 70 else year + 1900


def _parse_dmy_regex(text: str) -> date | None:
    """Parse DD.MM.YYYY (and / - , separators) after sanitization."""
    head = text.split(" ", 1)[0]
    match = _DATE_DMY_RE.match(head)
    if not match:
        return None
    day = int(match.group("d"))
    month = int(match.group("m"))
    year = _expand_two_digit_year(int(match.group("y")))
    try:
        return date(year, month, day)
    except ValueError:
        return None


def _parse_string_serial(text: str) -> date | None:
    """Text cells that contain only an Excel serial (General-formatted dates)."""
    cleaned = text.strip().replace(" ", "")
    if not cleaned.isdigit():
        return None
    return _excel_serial_to_date(float(cleaned))


def _excel_serial_to_date(value: float) -> date | None:
    """Convert an Excel serial date (days since 1899-12-30) to a `date`.

    Returns None for obviously-out-of-range values so we don't write garbage
    like year 4017 into the DB.
    """
    if value <= 0 or value > 80000:  # 80000 ≈ year 4118 — well beyond any real txn
        return None
    try:
        return (_EXCEL_EPOCH + timedelta(days=float(value))).date()
    except (OverflowError, ValueError):
        return None


def _parse_date(value: Any) -> date | None:
    """Best-effort date parser for bank Excel cells.

    Accepts: datetime / date objects, ISO/EU formatted strings (with or without
    a trailing time), and raw Excel serial numbers (floats/ints emitted when a
    cell is formatted as 'General' instead of 'Date'). Returns None on failure
    and logs a warning so the upload service can surface a row count to the UI.
    """
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, bool):  # bool is a subclass of int — reject early
        return None
    if isinstance(value, (int, float)):
        result = _excel_serial_to_date(float(value))
        if result is None:
            logger.warning(
                "Could not parse date from numeric cell: %r (%s)",
                value, type(value).__name__,
            )
        return result

    text = _sanitize_date_text(str(value))
    if not text:
        return None

    # Try the full set of accepted string formats first.
    for fmt in _DATE_STRING_FORMATS:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue

    # Some banks ship dates with extra trailing junk (timezone abbreviation,
    # weekday, etc.). Strip everything after the first space and retry the
    # date-only formats so e.g. "25.02.2026 Mo" still parses.
    head = text.split(" ", 1)[0]
    if head and head != text:
        for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%y", "%d/%m/%y"):
            try:
                return datetime.strptime(head, fmt).date()
            except ValueError:
                continue

    # Plain-text EU dates (apostrophe-prefixed, comma separators, etc.).
    parsed = _parse_dmy_regex(text)
    if parsed is not None:
        return parsed

    # Serial stored as text when the column is formatted General/@.
    parsed = _parse_string_serial(text)
    if parsed is not None:
        return parsed

    logger.warning(
        "Could not parse date from string cell: %r (%s)",
        text, type(value).__name__,
    )
    return None


def _is_summary_or_footer(row_values: list[Any]) -> bool:
    combined = " ".join(_normalize_header(v) for v in row_values if v is not None)
    return any(marker in combined for marker in STOP_MARKERS)


def _is_empty_transaction(row: dict[str, Any]) -> bool:
    return (
        row.get("transaction_date") is None
        and row.get("debited_amount") is None
        and row.get("credited_amount") is None
        and not row.get("comment")
    )


def _transaction_dedupe_key(row: ParsedBankRow) -> tuple:
    return (
        row.transaction_date,
        row.debited_amount,
        row.credited_amount,
        (row.comment or "").strip().casefold(),
        (row.transaction_type or "").strip().casefold(),
    )


def dedupe_parsed_rows(rows: list[ParsedBankRow]) -> tuple[list[ParsedBankRow], int]:
    """Remove duplicate transaction rows within a single statement upload."""
    seen: set[tuple] = set()
    unique: list[ParsedBankRow] = []
    skipped = 0
    for row in rows:
        key = _transaction_dedupe_key(row)
        if key in seen:
            skipped += 1
            continue
        seen.add(key)
        unique.append(row)
    return unique, skipped


_FILENAME_DATE_PATTERNS: tuple[tuple[re.Pattern[str], str], ...] = (
    (re.compile(r"(\d{4})(\d{2})(\d{2})"), "%Y%m%d"),
    (re.compile(r"(\d{4})[-_](\d{2})[-_](\d{2})"), "%Y-%m-%d"),
    (re.compile(r"(\d{4})[-_](\d{2})(?:[^0-9]|$)"), "%Y-%m"),
)


def extract_statement_date(
    filename: str,
    rows: list[ParsedBankRow],
) -> date:
    """Derive the statement period date from filename or transaction rows."""
    name = Path(filename).stem
    for pattern, fmt in _FILENAME_DATE_PATTERNS:
        match = pattern.search(name)
        if not match:
            continue
        if fmt == "%Y-%m":
            year, month = match.group(1), match.group(2)
            text = f"{year}-{month}-01"
            try:
                return datetime.strptime(text, "%Y-%m-%d").date()
            except ValueError:
                continue
        text = match.group(0).replace("_", "-")
        for parse_fmt in ("%Y%m%d", "%Y-%m-%d"):
            try:
                return datetime.strptime(text, parse_fmt).date()
            except ValueError:
                continue

    txn_dates = [r.transaction_date for r in rows if r.transaction_date is not None]
    if txn_dates:
        return max(txn_dates)

    raise ExcelParseError(
        "Could not determine statement date from the file name or transaction rows."
    )


def statement_id_from_date(statement_date: date) -> int:
    """Business statement ID shown in the UI (YYYYMMDD)."""
    return int(statement_date.strftime("%Y%m%d"))


def _map_columns(header_row: list[Any]) -> dict[str, int]:
    col_map: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        h = _normalize_header(cell)
        if not h:
            continue
        if any(k in h for k in REQUIRED_HEADERS["date"]) and "date" not in col_map:
            col_map["date"] = idx
        if any(k in h for k in REQUIRED_HEADERS["comment"]) and "comment" not in col_map:
            col_map["comment"] = idx
        if "debit" in h and "debited" not in col_map:
            col_map["debited"] = idx
        if "credit" in h and "credited" not in col_map:
            col_map["credited"] = idx
        if ("tipi" in h or "type" in h) and "transaction_type" not in col_map:
            col_map["transaction_type"] = idx
    return col_map


@debug_trace
def _find_header_row(rows: list[list[Any]]) -> tuple[int, dict[str, int]]:
    limit = min(HEADER_SCAN_ROWS, len(rows))
    for i in range(limit):
        row = rows[i]
        col_map = _map_columns(row)
        if "date" in col_map and "comment" in col_map:
            return i, col_map
    raise ExcelParseError(
        "Could not find bank statement headers (Data / Date, Komenti / Comment)."
    )


def _parse_row(row: list[Any], col_map: dict[str, int]) -> dict[str, Any]:
    def get(key: str) -> Any:
        idx = col_map.get(key)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    comment = _cell_str(get("comment"))
    return {
        "transaction_date": _parse_date(get("date")),
        "debited_amount": _parse_amount(get("debited")),
        "credited_amount": _parse_amount(get("credited")),
        "transaction_type": _cell_str(get("transaction_type")),
        "comment": comment,
        "detected_invoice_numbers": extract_invoice_numbers(comment),
    }


def _load_rows_xlsx(data: bytes) -> list[list[Any]]:
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        wb.close()
        raise ExcelParseError("Excel workbook has no active sheet.")
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


def _load_rows_xls(data: bytes) -> list[list[Any]]:
    import xlrd

    book = xlrd.open_workbook(file_contents=data)
    sheet = book.sheet_by_index(0)
    datemode = book.datemode
    rows: list[list[Any]] = []
    for r in range(sheet.nrows):
        row_values: list[Any] = []
        for c in range(sheet.ncols):
            cell = sheet.cell(r, c)
            # XL_CELL_DATE = 3. xlrd otherwise returns date cells as floats,
            # which `_parse_date` would only partially recover via the Excel
            # serial branch. Convert explicitly so downstream sees a datetime.
            if cell.ctype == xlrd.XL_CELL_DATE:
                try:
                    row_values.append(
                        xlrd.xldate.xldate_as_datetime(cell.value, datemode)
                    )
                    continue
                except Exception:
                    pass
            row_values.append(cell.value)
        rows.append(row_values)
    return rows


@debug_trace
def parse_bank_statement_excel(
    data: bytes,
    filename: str,
) -> list[ParsedBankRow]:
    ext = Path(filename).suffix.lower()
    if ext == ".xls":
        rows = _load_rows_xls(data)
    elif ext in (".xlsx", ".xlsm"):
        rows = _load_rows_xlsx(data)
    else:
        raise ExcelParseError("Unsupported file type. Upload .xlsx or .xls.")

    if not rows:
        raise ExcelParseError("No data rows found in the file.")

    header_idx, col_map = _find_header_row(rows)
    logger.debug(
        "Bank Excel header at row %d, col_map=%r (%s)",
        header_idx, col_map, type(col_map).__name__,
    )
    parsed: list[ParsedBankRow] = []

    for row in rows[header_idx + 1 :]:
        if _is_summary_or_footer(row):
            break
        row_dict = _parse_row(row, col_map)
        if _is_empty_transaction(row_dict):
            continue
        logger.debug(
            "  bank row parsed: date=%r (%s) debited=%r (%s) credited=%r (%s) "
            "type=%r (%s) detected=%r (%s)",
            row_dict["transaction_date"], type(row_dict["transaction_date"]).__name__,
            row_dict["debited_amount"], type(row_dict["debited_amount"]).__name__,
            row_dict["credited_amount"], type(row_dict["credited_amount"]).__name__,
            row_dict["transaction_type"], type(row_dict["transaction_type"]).__name__,
            row_dict["detected_invoice_numbers"],
            type(row_dict["detected_invoice_numbers"]).__name__,
        )
        parsed.append(
            ParsedBankRow(
                transaction_date=row_dict["transaction_date"],
                debited_amount=row_dict["debited_amount"],
                credited_amount=row_dict["credited_amount"],
                transaction_type=row_dict["transaction_type"],
                comment=row_dict["comment"],
                detected_invoice_numbers=row_dict["detected_invoice_numbers"],
            )
        )

    if not parsed:
        raise ExcelParseError("No transaction rows found.")

    return parsed
